import sys
import time

from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials
from dotenv import load_dotenv
from openpyxl.utils import column_index_from_string, get_column_letter
import os

load_dotenv()

token, sheet_id = os.getenv("TOKEN"), os.getenv("SHEET_ID")
SCOPES, SERVICE_ACCOUNT_FILE = ['https://www.googleapis.com/auth/spreadsheets'], 'google_auth.json'
credentials = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
gs_service = build("sheets", "v4", credentials=credentials)

firefox_options = Options()
firefox_options.add_argument("--headless")
driver = webdriver.Firefox(service=Service(), options=firefox_options)


def update_values(range_name, value_input_option, values, commendation_page_name):
    body = {'values': values}
    result = gs_service.spreadsheets().values().update(
        spreadsheetId=sheet_id, range=range_name, valueInputOption=value_input_option, body=body).execute()
    if result:
        print(f"{result.get('updatedCells')} cells of {commendation_page_name.capitalize()} updated.")
    return result


def open_page(page):
    try:
        url = f"https://www.seaofthieves.com/profile/reputation/{page}"
        driver.get(url)
        driver.add_cookie({"name": "rat", "value": token})
        driver.get(url)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.CLASS_NAME, "emblem-item__progress-text")))
        WebDriverWait(driver, 5)
        return driver.page_source
    except WebDriverException as e:
        print(f"Error opening page {page}: {e}.\nProb incorrect RATTE cookie.\n")
        return None
    except Exception as e:
        print(f"Unexpected error opening page {page}: {e}")
        return None


def get_data(commendation_page_name, page_source):
    try:
        if not page_source:
            return None
        return str(page_source.encode("utf-8"))
    except WebDriverException as e:
        print(f"Error scraping {commendation_page_name}: {e}. Restarting WebDriver...")
        driver.quit()
        return get_data(commendation_page_name, page_source)
    except Exception as e:
        print(f"Error scraping {commendation_page_name}: {e}")
        return None


def find_achievements(src, indices):
    result = []
    while "<div class=\"emblem-item__progress-text\">" in src:
        divindex = src.find("<div class=\"emblem-item__progress-text\">")
        src = src[divindex + 40:]
        result.append(src[:src.find("/")])
    return [result[i] for i in indices if i < len(result)]


def get_list(commendation_page_name, indices, page_source):
    print(f"Updating {commendation_page_name.capitalize()}")
    commendation_list = find_achievements(get_data(commendation_page_name, page_source), indices)
    if not commendation_list:
        print(f"Retrying {commendation_page_name} with a longer timeout.")
        commendation_list = find_achievements(get_data(commendation_page_name, page_source), indices)
        if not commendation_list:
            print(f"Failed to update {commendation_page_name}. Skipping.")
            return []
    return [[''] if commendation == '0' else [commendation] for commendation in commendation_list]


update_functions = {
    # "Commodation Page Name": ("Cell Range", [Commodation Indexs], "Commodation Base Page Name")
    "splashtails": ("D4:D8", [0, 1, 2, 3, 4], "HuntersCall"),
    "plentifins": ("J4:J8", [0, 1, 2, 3, 4], "HuntersCall"),
    "ancientscales": ("P4:P8", [0, 1, 2, 3, 4], "HuntersCall"),
    "wildsplashes": ("D12:D16", [0, 1, 2, 3, 4], "HuntersCall"),
    "devilfishes": ("J12:J16", [0, 1, 2, 3, 4], "HuntersCall"),
    "islehoppers": ("P12:P16", [0, 1, 2, 3, 4], "HuntersCall"),
    "pondies": ("D20:D24", [0, 1, 2, 3, 4], "HuntersCall"),
    "battlegills": ("J20:J24", [0, 1, 2, 3, 4], "HuntersCall"),
    "stormfishes": ("P20:P24", [0, 1, 2, 3, 4], "HuntersCall"),
    "wreckers": ("V4:V8", [0, 1, 2, 3, 4], "HuntersCall"),
    "cooking": ("AD5:BAD10", [0, 1, 2, 3, 4, 5], "HuntersCall"),
    "merrick's-accolades": ("AD13:AD14", [0, 7], "HuntersCall"),
    "shrouded-spoils": ("AD18:AD23", [2, 3, 4, 5, 6, 7], "BilgeRats"),
}


def convert_to_relative_ranges(start_cell, ranges):
    # The starting cell is assumed to be A1 (row 1, column 1)
    start_col = ''.join([c for c in start_cell if c.isalpha()])
    start_row = int(''.join([c for c in start_cell if c.isdigit()]))

    # Calculate the offset from A1 (which is (1, 1))
    col_offset = column_index_from_string(start_col) - 1  # Subtract 1 to make it relative
    row_offset = start_row - 1  # Subtract 1 to make it relative

    relative_ranges = {}

    for name, (cell_range, indices, base_page_name) in ranges.items():
        # Split the range into start and end cells (e.g., "E5:E9" -> "E5", "E9")
        start_cell, end_cell = cell_range.split(":")

        # Extract the row and column from the start and end cells
        start_col_idx = column_index_from_string(''.join([c for c in start_cell if c.isalpha()]))
        start_row_idx = int(''.join([c for c in start_cell if c.isdigit()]))
        end_col_idx = column_index_from_string(''.join([c for c in end_cell if c.isalpha()]))
        end_row_idx = int(''.join([c for c in end_cell if c.isdigit()]))

        # Shift the start and end coordinates by the calculated offsets
        relative_start_col = get_column_letter(start_col_idx + col_offset)
        relative_start_row = start_row_idx + row_offset
        relative_end_col = get_column_letter(end_col_idx + col_offset)
        relative_end_row = end_row_idx + row_offset

        # Store the updated range in a dictionary
        relative_ranges[name] = (
            f"{relative_start_col}{relative_start_row}:{relative_end_col}{relative_end_row}",
            indices,
            base_page_name
        )

    return relative_ranges


def get_starting_cell():
    name_map = {
        "ethan": "B30", "e": "B30",
        "chase": "B2", "c": "B2"
    }
    name = os.getenv('NAME').lower()

    if name in name_map:
        return name_map[name]
    else:
        print("Invalid Name, please update .env file")
        exit_app()


def exit_app():
    print("Exiting Now...")
    time.sleep(1)
    sys.exit()

starting_cell = get_starting_cell()
print("Beginning Loading Information...\n")

relative_update_functions = convert_to_relative_ranges(starting_cell, update_functions)
for commendation_page_name, (cell_range, indices, base_page_name) in relative_update_functions.items():
    page_source = open_page(f"{base_page_name}/{commendation_page_name}")
    commendation_list = get_list(commendation_page_name, indices, page_source)
    if commendation_list:
        update_values(cell_range, "USER_ENTERED", commendation_list, commendation_page_name)
exit_app()
